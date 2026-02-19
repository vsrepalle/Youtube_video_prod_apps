import os, json, subprocess, sys, time, msvcrt, socket
from datetime import datetime
from openai import OpenAI
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

load_dotenv()
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# --- 1. DEFINE LOGGER FIRST (Avoids NameError) ---
def log_to_excel(channel, topic, status="Success"):
    log_file = os.path.join(BASE_DIR, "TrendWave_Logs.xlsx")
    timestamp = datetime.now()
    log_time = timestamp.strftime("%Y-%m-%d %H:%M:%S")
    video_path = os.path.join(BASE_DIR, "Ready_to_Upload", f"Render_{channel}_{timestamp.strftime('%Y-%m-%d')}.mp4")
    hyperlink = f'=HYPERLINK("{video_path}", "Open Video")'

    if not os.path.exists(log_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Execution Logs"
        ws.append(["Timestamp", "Channel", "Topic", "Status", "Video Link"])
    else:
        wb = load_workbook(log_file)
        ws = wb.active

    ws.append([log_time, channel, topic, status, hyperlink])
    ws.cell(row=ws.max_row, column=5).font = Font(color="0000FF", underline="single")
    wb.save(log_file)

# --- 2. DEFINE AI LOGIC WITH CLEANING ---
def get_ai_answer(channel, user_topic):
    client = OpenAI(base_url="https://api.groq.com/openai/v1", api_key=os.getenv("GROQ_API_KEY"))
    prompt = f"Return ONLY a JSON array for {channel} news about {user_topic}. No intro text. Use strict JSON format."
    
    try:
        completion = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1
        )
        raw_content = completion.choices[0].message.content
        
        # CLEANER: Extract only what is between [ and ] to prevent JSONDecodeError
        start_index = raw_content.find("[")
        end_index = raw_content.rfind("]") + 1
        if start_index != -1 and end_index != 0:
            return raw_content[start_index:end_index]
        return raw_content
    except Exception as e:
        return f"Error: {e}"

# --- 3. DEFINE RUNNER ---
def save_and_run(json_data, channel_name, topic):
    try:
        folder = os.path.join(BASE_DIR, "generated_content")
        os.makedirs(folder, exist_ok=True)
        static_filepath = os.path.join(folder, "news_data.json")
        
        with open(static_filepath, "w", encoding="utf-8") as f:
            f.write(json_data)
        
        controller_path = os.path.join(BASE_DIR, "controller.py")
        print(f"🚀 Triggering Controller for {channel_name}...")
        subprocess.run(["python", controller_path, static_filepath], check=True)
        log_to_excel(channel_name, topic, "Success")
    except Exception as e:
        log_to_excel(channel_name, topic, f"Error: {str(e)}")
        print(f"❌ Execution failed: {e}")

# --- 4. MAIN SCHEDULER ---
def main():
    print("--- 🤖 GEMINI-CORE AUTONOMOUS SCHEDULER STARTED ---")
    last_run_time = ""
    
    # Load config from file
    with open('config.json', 'r') as f:
        config = json.load(f)

    while True:
        now = datetime.now()
        current_time = now.strftime("%H:%M")
        
        if current_time != last_run_time:
            # Check all channels in config
            for name, data in config['channels'].items():
                if current_time in data['schedule']:
                    print(f"\n⏰ SLOT MATCH: {current_time} for {name}")
                    
                    # 15s Override Loop (msvcrt for Windows)
                    print("Press ANY KEY for manual override...")
                    start_wait = time.time()
                    override = False
                    while time.time() - start_wait < 15:
                        if msvcrt.kbhit():
                            msvcrt.getch()
                            override = True
                            break
                        time.sleep(0.1)

                    if override:
                        topic = input("Enter custom topic: ")
                    else:
                        topic = data['prompt']
                    
                    result = get_ai_answer(name, topic)
                    save_and_run(result, name, topic)
                    last_run_time = current_time
        
        time.sleep(30)

if __name__ == "__main__":
    main()